import datetime
import time

from PyQt6.QtCore import QThreadPool, Qt, pyqtSignal
from PyQt6.QtWidgets import QWidget, QStatusBar, QGridLayout, QPushButton
from loguru import logger
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from app.threads.worker import Worker
from app.utils import close_excel, get_species, get_species_name_by_guid
from app.vetis.main import Vetis
from app.windows.log import Log


class Report(QWidget):

    def __init__(self, log_window: Log, status_bar: QStatusBar, parent=None):
        super().__init__(parent)

        self.allowed_districts: list = []
        self.log_window = log_window
        self.status_bar = status_bar
        self.thread_pool = QThreadPool()

        self.layout = QGridLayout()
        self.layout.setVerticalSpacing(10)
        self.layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.setLayout(self.layout)

        # ToDo start button
        self.start_button = QPushButton()
        self.start_button.setText("Получить отчет")
        self.start_button.setToolTip("Получение отчета о зарегистрированных животных")
        # self.start_button.setEnabled(False)
        self.start_button.clicked.connect(self.get_report)
        self.layout.addWidget(self.start_button, 0, 1, 1, 1)

    def get_report(self):
        worker = Worker(self.get_report_thread)
        worker.signals.progress.connect(self.report_progress)
        self.thread_pool.start(worker)

    def get_report_thread(self, progress_callback: pyqtSignal):
        keeping_purposes: list = Vetis.dictionary_service.get_animal_keeping_purpose_list()
        specieses: list = Vetis.dictionary_service.get_animal_species_list()
        breeds: list = Vetis.dictionary_service.get_animal_breed_list()
        keeping_types: list = Vetis.dictionary_service.get_animal_keeping_type_list()
        marking_locations: list = Vetis.dictionary_service.get_animal_marking_location_list()
        registred_animals = []
        st = datetime.datetime.now()
        start_date = datetime.date(year=2022, month=1, day=1)
        while start_date < datetime.datetime.now().date():
            if len(registred_animals) >= 5000:
                self.save_animals_to_excel(registred_animals)
                registred_animals.clear()
            logger.debug(len(registred_animals))
            progress_callback.emit(str(len(registred_animals)))
            logger.info(start_date)
            for species in specieses:
                progress_callback.emit(str(species["name"]))
                logger.warning(species["name"])
                offset = 0

                while True:
                    st2 = datetime.datetime.now()
                    try:
                        Vetis.herriot.get_animal_registration_changes_list(
                            start_date=start_date,
                            end_date=start_date + datetime.timedelta(days=30),
                            animal_species=species["guid"],
                            offset=offset
                        )
                        response = Vetis.herriot.get_finished_response()
                    except Exception as error:
                        logger.error(error)
                        time.sleep(1)
                        continue
                    animals = response["result"]["_value_1"]["animalRegistrationList"]["animalRegistration"]
                    # lasting = [animal["last"] for animal in animals]
                    # logger.info(
                    #     [[label["attachedLabel"]["animalID"]["_value_1"] for label in animal["specifiedAnimalIdentity"]]
                    #      for animal in animals])
                    progress_callback.emit(str(len(animals)))
                    logger.info(len(animals))
                    for animal in animals:
                        if animal["last"]:
                            registred_animals.append(animal)
                    if response["result"]["_value_1"]["animalRegistrationList"]["hasMore"]:
                        offset += 100
                    else:
                        break
                    progress_callback.emit(str(datetime.datetime.now() - st2))
                    logger.debug(datetime.datetime.now() - st2)
            start_date += datetime.timedelta(days=30)
        self.save_animals_to_excel(registred_animals)
        progress_callback.emit("Почти-почти")
        progress_callback.emit(str(datetime.datetime.now() - st))
        logger.debug(datetime.datetime.now() - st)

    def save_animals_to_excel(self, registred_animals: list):
        specieses: list = Vetis.dictionary_service.get_animal_species_list()
        wb = Workbook()
        ws = wb[wb.sheetnames[0]]
        ws.title = "Учтенные животные"
        ws.cell(row=1, column=1, value="Дата/время").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Владелец").font = Font(bold=True)
        ws.cell(row=1, column=3, value="№ ПО").font = Font(bold=True)
        ws.cell(row=1, column=4, value="Регион").font = Font(bold=True)
        ws.cell(row=1, column=5, value="Район").font = Font(bold=True)
        ws.cell(row=1, column=6, value="Нас. пункт").font = Font(bold=True)
        ws.cell(row=1, column=7, value="Полный адрес").font = Font(bold=True)
        ws.cell(row=1, column=8, value="Инд./групп.").font = Font(bold=True)
        ws.cell(row=1, column=9, value="Вид животного").font = Font(bold=True)
        ws.cell(row=1, column=10, value="Инд. номер").font = Font(bold=True)
        ws.cell(row=1, column=11, value="Рег. номер").font = Font(bold=True)
        ws.cell(row=1, column=12, value="GUID").font = Font(bold=True)
        for animal in registred_animals:
            if animal["registrationStatus"]["_value_1"] != "ACTIVE":
                continue
            response_so = Vetis.enterprise_service.get_supervised_object_by_guid(
                animal["keepingDetails"]["operatorSupervisedObject"]["guid"]
            )
            row_idx = ws.max_row + 1
            ws.cell(
                row=row_idx,
                column=1,
                value=animal.updateDate.strftime("%d.%m.%Y %H:%M"),
            )
            ws.cell(row=row_idx, column=2, value=response_so.businessEntity.name)
            ws.cell(row=row_idx, column=3, value=animal["keepingDetails"]["operatorSupervisedObject"]["approvalNumber"])
            ws.cell(row=row_idx, column=4, value=response_so.enterprise.address.region.name)
            if response_so.enterprise.address.district is not None:
                ws.cell(
                    row=row_idx, column=5, value=response_so.enterprise.address.district.name
                )
            ws.cell(
                row=row_idx, column=6, value=response_so.enterprise.address.locality.name
            )
            ws.cell(row=row_idx, column=7, value=response_so.enterprise.address.addressView)
            if animal["identityType"]["_value_1"] == "INDIVIDUAL":
                ws.cell(row=row_idx, column=8, value="Индивидуальный")
                ws.cell(
                    row=row_idx, column=9,
                    value=get_species_name_by_guid(animal["specifiedAnimal"]["species"]["guid"], specieses)
                )
            elif animal["identityType"]["_value_1"] == "GROUP":
                ws.cell(row=row_idx, column=8, value="Групповой")
                ws.cell(
                    row=row_idx, column=9,
                    value=get_species_name_by_guid(animal["specifiedAnimalGroup"]["species"]["guid"], specieses)
                )
            if animal["specifiedAnimalIdentity"] and len(animal["specifiedAnimalIdentity"]) > 0:
                try:
                    ws.cell(
                        row=row_idx,
                        column=10,
                        value=animal["specifiedAnimalIdentity"][0]["attachedLabel"]["animalID"]["_value_1"]
                    )
                except Exception as error:
                    logger.error(f"{type(error)} {error}")
            ws.cell(
                row=row_idx, column=11, value=animal["registrationNumber"]["_value_1"]
            )
            ws.cell(row=row_idx, column=12, value=animal["guid"])
        try:
            wb.save(f"Зарегистрированные животные свод {datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx")
        except PermissionError:
            close_excel()
            wb.save(f"Зарегистрированные животные свод {datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx")
        wb.close()

    def report_progress(self, info: str):
        self.log_window.write(f"{info}")
