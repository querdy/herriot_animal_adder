import time
from datetime import datetime
from pathlib import Path
from random import randint

import pandas as pd
from PyQt6.QtCore import QThreadPool, Qt, QMutex, pyqtSignal
from PyQt6.QtWidgets import (
    QWidget,
    QStatusBar,
    QPushButton,
    QGridLayout,
    QFileDialog,
    QTableWidget,
    QTableWidgetItem,
    QFrame,
    QLabel,
    QCheckBox,
)
from loguru import logger
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from pandas import Timestamp

from app.threads.worker import Worker
from app.utils import (
    check_species,
    check_breed,
    check_keeping_type,
    check_marking_location,
    check_keeping_purpose,
    get_keeping_type,
    get_species,
    get_breed,
    get_supervised_object,
    get_marking_location,
    get_keeping_purpose,
    check_animal_gender,
    check_breeding_value,
    check_initial_identification,
    check_marking_means,
    check_label_id,
    check_supervised_object,
    check_keeping_place_allowed, close_excel,
)
from app.vetis.main import Vetis
from app.vetis.schemas.enums import (
    InitialIdentificationType,
    AnimalBreedingValueType,
    AnimalGender,
    AnimalLabelType,
    MarkingMeansType,
    RegistrationStatus,
)
from app.vetis.schemas.herriot import ComplexDate, AnimalRegistration
from app.vetis.schema_builders import get_animal_registration_ame_single
from app.windows.base import CustomQTableWidget
from app.windows.log import Log


class IndividualRegister(QWidget):
    stop_workers = pyqtSignal()

    def __init__(self, log_window: Log, status_bar: QStatusBar, parent=None):
        super().__init__(parent)

        self.allowed_districts: list = []
        self.log_window = log_window
        self.status_bar = status_bar
        self.thread_pool = QThreadPool()
        self.save_excel_mutex = QMutex()

        parent.signals.allowed_districts_received.connect(
            self.allowed_districts_received
        )

        self.layout = QGridLayout()
        self.layout.setVerticalSpacing(10)
        self.layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.setLayout(self.layout)

        self.animals_for_register: list[tuple[int, AnimalRegistration]] = []

        # ToDo select file button
        self.select_file_button = QPushButton()
        self.select_file_button.setText("Выбрать файл")
        self.select_file_button.clicked.connect(self.open_file_dialog)
        self.layout.addWidget(self.select_file_button, 0, 0, 1, 1)

        # ToDo start button
        self.start_button = QPushButton()
        self.start_button.setText("Начать учет")
        self.start_button.setToolTip("Внести данные из выбранной таблицы")
        self.start_button.setEnabled(False)
        self.start_button.clicked.connect(self.register_animal)
        self.layout.addWidget(self.start_button, 0, 1, 1, 1)

        # ToDo stop button
        self.stop_button = QPushButton()
        self.stop_button.setText("Остановить")
        self.stop_button.setEnabled(False)
        self.stop_button.clicked.connect(self.stop_all_thread)
        self.layout.addWidget(self.stop_button, 0, 2, 1, 1)

        # ToDo is delay checkbox
        self.delay_checkbox = QCheckBox("Имитация бурной деятельности")
        self.delay_checkbox.setToolTip(
            "Случайная задержка 45-90 сек.\nЭто примерное время внесения через веб-версию"
        )
        self.delay_checkbox.setChecked(False)
        self.layout.addWidget(self.delay_checkbox, 0, 3, 1, 1)

        # ToDo selected file label
        self.selected_file_label = QLabel("Выбранный файл:")
        self.layout.addWidget(self.selected_file_label, 1, 0, 1, 1)

        # ToDo selected file
        self.selected_file_label = QLabel()
        self.layout.addWidget(self.selected_file_label, 1, 1, 1, 6)

        # ToDo main table
        self.table = CustomQTableWidget()
        self.table.setFrameShape(QFrame.Shape.NoFrame)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_data: pd.DataFrame = pd.DataFrame()
        self.layout.addWidget(self.table, 2, 0, 1, 10)

    def stop_all_thread(self):
        self.stop_workers.emit()

    def allowed_districts_received(self, districts: list[str]):
        self.allowed_districts = districts

    def read_excel_thread(self, path: Path) -> None:
        table_data = pd.read_excel(path)
        table_data.dropna(axis='index', how='all', inplace=True)
        self.table_data = table_data.map(lambda x: x.strip() if isinstance(x, str) else x)
        self.table.set_row_and_column_count(self.table_data.shape)
        self.table.setHorizontalHeaderLabels(self.table_data.columns.values)
        for row_idx, row_data in self.table_data.iterrows():
            for col_idx, value in enumerate(row_data):
                if pd.notna(value):
                    if isinstance(value, Timestamp):
                        self.table.setItem(
                            row_idx,
                            col_idx,
                            QTableWidgetItem(value.date().strftime("%d.%m.%Y")),
                        )
                    else:
                        self.table.setItem(
                            row_idx, col_idx, QTableWidgetItem(str(value))
                        )
                else:
                    self.table.setItem(row_idx, col_idx, QTableWidgetItem())
        self.table.resizeColumnsToContents()

    def read_excel(self, path: Path) -> None:
        worker = Worker(self.read_excel_thread, path)
        worker.signals.finished.connect(self.read_excel_finished)
        self.thread_pool.start(worker)

    def read_excel_finished(self) -> None:
        self.check_table_for_correctness()

    def open_file_dialog(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Выберете таблицу для учета", filter="Excel (*.xlsx)"
        )
        if filename:
            self.selected_file_label.setText(filename)
            self.read_excel(Path(filename))

    def register_animal(self):
        try:
            is_work_imitation = False
            if self.delay_checkbox.checkState() == Qt.CheckState.Checked:
                is_work_imitation = True
            worker = Worker(self.register_animal_thread, is_work_imitation)
            self.stop_workers.connect(worker.stop)
            worker.signals.progress.connect(self.report_progress)
            worker.signals.finished.connect(lambda: self.stop_button.setEnabled(False))
            worker.signals.finished.connect(
                lambda: self.select_file_button.setEnabled(True)
            )
            worker.signals.finished.connect(lambda: self.delay_checkbox.setEnabled(True))
            self.delay_checkbox.setEnabled(False)
            self.start_button.setEnabled(False)
            self.select_file_button.setEnabled(False)
            self.stop_button.setEnabled(True)
            self.thread_pool.start(worker)
        except Exception as err:
            logger.warning(f"{type(err)} {err}")

    def register_animal_thread(self, is_work_imitation: bool, progress_callback: pyqtSignal, worker_object: Worker):
        progress_callback.emit(f"Внесение данных в хорриот: начало")
        for animal in self.animals_for_register:
            if worker_object.is_stop:
                progress_callback.emit(f"Учёт остановлен")
                self.start_button.setEnabled(True)
                self.status_bar.clearMessage()
                return
            if is_work_imitation:
                sleep_seconds = randint(45, 90)
                progress_callback.emit(f"Имитирую бурную деятельность: {sleep_seconds} сек.")
                time.sleep(sleep_seconds)
            self.table.selectRow(animal[0])
            self.status_bar.showMessage(f"Отправка запроса. Строка {animal[0] + 1}")
            application = Vetis.herriot.register_animal(animal[1])
            response = Vetis.herriot.get_finished_response(application)
            if response.status == "REJECTED":
                self.table.set_color_to_row(animal[0], (255, 220, 220))
                progress_callback.emit(f"Ошибка отправки запроса в строке {animal[0] + 1}:")
                for error in response.errors.error:
                    progress_callback.emit(f"{error}")
            else:
                self.table.set_color_to_row(animal[0], (220, 255, 220))
                self.save_new_animal_to_excel(
                    response["result"]["_value_1"]["animalRegistration"][0]["guid"],
                    self.save_excel_mutex,
                )
        self.animals_for_register = []
        self.table.clearSelection()
        self.status_bar.clearMessage()
        progress_callback.emit(f"Внесение данных в хорриот: конец")

    def save_new_animal_to_excel(self, animal_guid: str, mutex: QMutex):
        worker = Worker(self.save_new_animal_to_excel_thread, animal_guid, mutex)
        worker.signals.progress.connect(self.report_progress)
        self.thread_pool.start(worker)

    @staticmethod
    def save_new_animal_to_excel_thread(animal_guid: str, mutex: QMutex, progress_callback, worker_object: Worker):
        application = Vetis.herriot.get_animal_registration_by_guid(animal_guid)
        response = Vetis.herriot.get_finished_response(application)
        if response.status == "REJECTED":
            for error in response.errors.error:
                progress_callback.emit(f"{error}")
            return
        registred_animal = response["result"]["_value_1"]["animalRegistration"]
        response_so = Vetis.enterprise_service.get_supervised_object_by_guid(
            registred_animal["keepingDetails"]["operatorSupervisedObject"]["guid"]
        )
        mutex.lock()
        try:
            wb = load_workbook("Зарегистрированные животные.xlsx", data_only=True)
            ws = wb[wb.sheetnames[0]]
            if ws.max_row >= 5000:
                timestamp_for_file = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
                wb.save(f"Зарегистрированные животные_{timestamp_for_file}.xlsx")
                progress_callback.emit(f"Зарегистрированные животные_{timestamp_for_file}.xlsx сохранен")
                raise FileNotFoundError("Файл сохранен с новым именем. Необходимо создать новый")
        except FileNotFoundError:
            wb = Workbook()
            ws = wb[wb.sheetnames[0]]
            ws.title = "Учтенные животные"
            ws.cell(row=1, column=1, value="Дата/время").font = Font(bold=True)
            ws.cell(row=1, column=2, value="Владелец").font = Font(bold=True)
            ws.cell(row=1, column=3, value="Регион").font = Font(bold=True)
            ws.cell(row=1, column=4, value="Район").font = Font(bold=True)
            ws.cell(row=1, column=5, value="Нас. пункт").font = Font(bold=True)
            ws.cell(row=1, column=6, value="Полный адрес").font = Font(bold=True)
            ws.cell(row=1, column=7, value="Вид животного").font = Font(bold=True)
            ws.cell(row=1, column=8, value="Инд. номер").font = Font(bold=True)
            ws.cell(row=1, column=9, value="GUID").font = Font(bold=True)
            ws.cell(row=1, column=10, value="Рег. номер").font = Font(bold=True)
        row_idx = ws.max_row + 1
        ws.cell(
            row=row_idx,
            column=1,
            value=registred_animal.updateDate.strftime("%d.%m.%Y %H:%M"),
        )
        ws.cell(row=row_idx, column=2, value=response_so.businessEntity.name)
        ws.cell(row=row_idx, column=3, value=response_so.enterprise.address.region.name)
        ws.cell(
            row=row_idx, column=4, value=response_so.enterprise.address.district.name
        )
        ws.cell(
            row=row_idx, column=5, value=response_so.enterprise.address.locality.name
        )
        ws.cell(row=row_idx, column=6, value=response_so.enterprise.address.addressView)
        ws.cell(
            row=row_idx, column=7, value=registred_animal.specifiedAnimal.species.name
        )
        ws.cell(
            row=row_idx,
            column=8,
            value=registred_animal["specifiedAnimalIdentity"][0]["attachedLabel"]["animalID"]["_value_1"]
        )
        ws.cell(row=row_idx, column=9, value=animal_guid)
        ws.cell(
            row=row_idx, column=10, value=registred_animal["registrationNumber"]["_value_1"]
        )
        try:
            wb.save("Зарегистрированные животные.xlsx")
        except PermissionError:
            close_excel()
            wb.save("Зарегистрированные животные.xlsx")
        wb.close()
        mutex.unlock()
        progress_callback.emit(
            f"Животное с номером "
            f"{registred_animal["specifiedAnimalIdentity"][0]["attachedLabel"]["animalID"]["_value_1"]} "
            f"сохранено (строка {str(row_idx)})"
        )

    def check_table_for_correctness(self):
        self.start_button.setEnabled(False)
        worker = Worker(self.check_table_for_correctness_thread)
        worker.signals.progress.connect(self.report_progress)
        worker.signals.finished.connect(lambda: self.start_button.setEnabled(True))
        self.thread_pool.start(worker)

    def check_table_for_correctness_thread(self, progress_callback: pyqtSignal, worker_object: Worker):
        progress_callback.emit(f"Получение справочников: начало")
        keeping_purpose_list: list = Vetis.dictionary_service.get_animal_keeping_purpose_list()
        species_list: list = Vetis.dictionary_service.get_animal_species_list()
        breed_list: list = Vetis.dictionary_service.get_animal_breed_list()
        keeping_type_list: list = Vetis.dictionary_service.get_animal_keeping_type_list()
        marking_location_list: list = Vetis.dictionary_service.get_animal_marking_location_list()
        progress_callback.emit(f"Получение справочников: конец")
        progress_callback.emit(f"Проверка правильности ввода данных в таблицу: начало")
        animals_for_register: list = []
        for row_idx, row_data in self.table_data.iterrows():
            self.status_bar.showMessage(f"Проверка строки {row_idx + 1}")
            is_valid = True
            row_animal: dict = {
                "species": row_data.iloc[0],
                "breed": row_data.iloc[1],
                "keeping_type": row_data.iloc[2],
                "birth_date": row_data.iloc[3],
                "gender": row_data.iloc[4],
                "breeding_value": row_data.iloc[5],
                "initial_identification": row_data.iloc[6],
                "actual_date_main": row_data.iloc[7],
                "marking_means_main": row_data.iloc[8],
                "marking_location_main": row_data.iloc[9],
                "label_id_main": row_data.iloc[10],
                "actual_date_additional_1": row_data.iloc[11],
                "marking_means_additional_1": row_data.iloc[12],
                "marking_location_additional_1": row_data.iloc[13],
                "label_id_additional_1": row_data.iloc[14],
                "actual_date_additional_2": row_data.iloc[15],
                "marking_means_additional_2": row_data.iloc[16],
                "marking_location_additional_2": row_data.iloc[17],
                "label_id_additional_2": row_data.iloc[18],
                "colour": row_data.iloc[19],
                "name": row_data.iloc[20],
                "keeping_place": row_data.iloc[21],
                "keeping_place_inn": row_data.iloc[22],
                "birth_place": row_data.iloc[23],
                "birth_place_inn": row_data.iloc[24],
                "keeping_purpose_1": row_data.iloc[25],
                "keeping_purpose_2": row_data.iloc[26],
                "keeping_purpose_3": row_data.iloc[27],
                "keeping_purpose_4": row_data.iloc[28],
                "keeping_purpose_5": row_data.iloc[29],
                "parent_1": row_data.iloc[30]
            }
            if not check_species(row_animal["species"], species_list):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[0]}"
                )
                is_valid = False

            if not check_breed(row_animal["breed"], row_animal["species"], breed_list):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[1]}"
                )
                is_valid = False

            if not check_keeping_type(row_animal["keeping_type"], keeping_type_list):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[2]}"
                )
                is_valid = False

            if isinstance(row_animal["birth_date"], Timestamp):
                row_animal["birth_date"] = row_animal["birth_date"].date()
                if row_animal["birth_date"] > datetime.now().date():
                    progress_callback.emit(
                        f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[3]}"
                    )
                    is_valid = False

            if not check_animal_gender(row_animal["gender"]):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[4]}"
                )
                is_valid = False

            if not check_breeding_value(row_animal["breeding_value"]):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[5]}"
                )
                is_valid = False

            if not check_initial_identification(row_animal["initial_identification"]):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[6]}"
                )
                is_valid = False

            if pd.notna(row_animal["actual_date_main"]):
                if row_animal["actual_date_main"].date() > datetime.now().date():
                    progress_callback.emit(
                        f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[7]}"
                    )
                    is_valid = False

            if not check_marking_means(row_animal["marking_means_main"]):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[8]}"
                )
                is_valid = False

            if not check_marking_location(row_animal["marking_location_main"], marking_location_list):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[9]}"
                )
                is_valid = False

            if not check_label_id(row_animal["label_id_main"], AnimalLabelType.MAIN.name):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[10]}"
                )
                is_valid = False

            if pd.notna(row_animal["actual_date_additional_1"]):
                if row_animal["actual_date_additional_1"].date() > datetime.now().date():
                    progress_callback.emit(
                        f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[11]}"
                    )
                    is_valid = False

            if not check_marking_means(row_animal["marking_means_additional_1"]):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[12]}"
                )
                is_valid = False

            if not check_marking_location(row_animal["marking_location_additional_1"], marking_location_list):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[13]}"
                )
                is_valid = False

            if not check_label_id(row_animal["label_id_additional_1"], AnimalLabelType.ADDITIONAL.name):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[14]}"
                )
                is_valid = False

            if pd.notna(row_animal["actual_date_additional_2"]):
                if row_animal["actual_date_additional_2"].date() > datetime.now().date():
                    progress_callback.emit(
                        f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[15]}"
                    )
                    is_valid = False

            if not check_marking_means(row_animal["marking_means_additional_2"]):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[16]}"
                )
                is_valid = False

            if not check_marking_location(row_animal["marking_location_additional_2"], marking_location_list):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[17]}"
                )
                is_valid = False

            if not check_label_id(row_animal["label_id_additional_2"], AnimalLabelType.ADDITIONAL.name):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[18]}"
                )
                is_valid = False

            if not check_supervised_object(
                    row_animal["keeping_place"],
                    Vetis.enterprise_service.get_be_supervised_object_list(
                        inn=row_animal["keeping_place_inn"]
                    ),
            ):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[21]}"
                    f" - указанный поднадзорный объект не существует или принадлежит другому владельцу (ИНН)"
                )
                is_valid = False

            keeping_place_so = get_supervised_object(
                row_animal["keeping_place"],
                Vetis.enterprise_service.get_be_supervised_object_list(
                    inn=row_animal["keeping_place_inn"]
                ),
            )
            if keeping_place_so is None:
                is_valid = False
            elif not check_keeping_place_allowed(keeping_place_so, self.allowed_districts):
                progress_callback.emit(
                    f"Строка {row_idx + 1} у вас нет доступа к учету животных из этого района"
                )
                is_valid = False
            if all([pd.notna(row_animal["birth_place"]), row_animal["birth_place_inn"]]):
                if not check_supervised_object(
                        row_animal["birth_place"],
                        Vetis.enterprise_service.get_be_supervised_object_list(
                            inn=row_animal["birth_place_inn"]
                        ),
                ):
                    progress_callback.emit(
                        f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[23]}"
                    )
                    is_valid = False

            if not check_keeping_purpose(
                    row_animal["keeping_purpose_1"], keeping_purpose_list
            ):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[25]}"
                )
                is_valid = False

            if pd.notna(row_animal["keeping_purpose_2"]) and not check_keeping_purpose(
                    row_animal["keeping_purpose_2"], keeping_purpose_list
            ):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[26]}"
                )
                is_valid = False

            if pd.notna(row_animal["keeping_purpose_3"]) and not check_keeping_purpose(
                    row_animal["keeping_purpose_3"], keeping_purpose_list
            ):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[27]}"
                )
                is_valid = False

            if pd.notna(row_animal["keeping_purpose_4"]) and not check_keeping_purpose(
                    row_animal["keeping_purpose_4"], keeping_purpose_list
            ):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[28]}"
                )
                is_valid = False

            if pd.notna(row_animal["keeping_purpose_5"]) and not check_keeping_purpose(
                    row_animal["keeping_purpose_5"], keeping_purpose_list
            ):
                progress_callback.emit(
                    f"Строка {row_idx + 1} ошибка в столбце {self.table_data.columns[29]}"
                )
                is_valid = False

            if is_valid:
                try:
                    animal = {
                        "initial_identification_type": InitialIdentificationType(
                            row_animal["initial_identification"]),
                        "breeding_value_type": AnimalBreedingValueType(row_animal["breeding_value"]),
                        "keeping_operator_supervised_object": get_supervised_object(
                            row_animal["keeping_place"],
                            Vetis.enterprise_service.get_be_supervised_object_list(
                                inn=row_animal["keeping_place_inn"]
                            ),
                        ).guid,
                        "keeping_type": get_keeping_type(
                            row_animal["keeping_type"], keeping_type_list
                        ),
                        "keeping_purpose": [
                            get_keeping_purpose(
                                purpose, keeping_purpose_list
                            )
                            for purpose in (
                                row_animal["keeping_purpose_1"],
                                row_animal["keeping_purpose_2"],
                                row_animal["keeping_purpose_3"],
                                row_animal["keeping_purpose_4"],
                                row_animal["keeping_purpose_5"],
                            )
                            if pd.notna(purpose)
                        ],
                        "species": get_species(row_animal["species"], species_list),
                        "breed": get_breed(row_animal["breed"], row_animal["species"], breed_list) if pd.notna(
                            row_animal["breed"]) else None,
                        "gender": AnimalGender(row_animal["gender"]),
                        "birth_date": (
                            ComplexDate(
                                year=row_animal["birth_date"].year,
                                month=row_animal["birth_date"].month,
                                day=row_animal["birth_date"].day,
                            )
                            if pd.notna(row_animal["birth_date"])
                            else None
                        ),
                        "colour": str(row_animal["colour"]) if pd.notna(row_animal["colour"]) else None,
                        "name": str(row_animal["name"]) if pd.notna(row_animal["name"]) else None,
                        "registration_status": RegistrationStatus.ACTIVE,
                    }
                    if all([pd.notna(row_animal["birth_place"]), row_animal["birth_place_inn"]]):
                        animal.update(
                            {
                                "birth_location": get_supervised_object(
                                    row_animal["birth_place"],
                                    Vetis.enterprise_service.get_be_supervised_object_list(
                                        inn=row_animal["birth_place_inn"]
                                    ),
                                )["guid"]
                            }
                        )
                    if pd.notna(row_animal["parent_1"]):
                        animal.update(
                            {
                                "pedigree_info_external": [row_animal["parent_1"]]
                            }
                        )
                    if all([pd.notna(item) for item in
                            (row_animal["marking_means_additional_2"], row_animal["actual_date_additional_2"],
                             row_animal["marking_location_additional_2"], row_animal["label_id_additional_2"])]):
                        animal.update(
                            {
                                "marking_means_additional_2": MarkingMeansType(
                                    row_animal["marking_means_additional_2"]),
                                "actual_date_additional_2": ComplexDate(
                                    year=row_animal["actual_date_additional_2"].year,
                                    month=row_animal["actual_date_additional_2"].month,
                                    day=row_animal["actual_date_additional_2"].day,
                                ),
                                "attachment_location_additional_2": get_marking_location(
                                    row_animal["marking_location_additional_2"], marking_location_list
                                ),
                                "animal_id_additional_2": str(row_animal["label_id_additional_2"]),
                                "ref_doc_uuid": str(row_animal["label_id_additional_2"]),
                            }
                        )
                    if all([pd.notna(item) for item in
                            (row_animal["marking_means_additional_1"], row_animal["actual_date_additional_1"],
                             row_animal["marking_location_additional_1"], row_animal["label_id_additional_1"])]):
                        animal.update(
                            {
                                "marking_means_additional_1": MarkingMeansType(
                                    row_animal["marking_means_additional_1"]),
                                "actual_date_additional_1": ComplexDate(
                                    year=row_animal["actual_date_additional_1"].year,
                                    month=row_animal["actual_date_additional_1"].month,
                                    day=row_animal["actual_date_additional_1"].day,
                                ),
                                "attachment_location_additional_1": get_marking_location(
                                    row_animal["marking_location_additional_1"], marking_location_list
                                ),
                                "animal_id_additional_1": str(row_animal["label_id_additional_1"]),
                                "ref_doc_uuid": str(row_animal["label_id_additional_1"]),
                            }
                        )
                    if all([pd.notna(item) for item in
                            (row_animal["marking_means_main"], row_animal["actual_date_main"],
                             row_animal["marking_location_main"], row_animal["label_id_main"])]):
                        animal.update(
                            {
                                "marking_means_main": MarkingMeansType(row_animal["marking_means_main"]),
                                "actual_date_main": ComplexDate(
                                    year=row_animal["actual_date_main"].year,
                                    month=row_animal["actual_date_main"].month,
                                    day=row_animal["actual_date_main"].day,
                                ),
                                "attachment_location_main": get_marking_location(
                                    row_animal["marking_location_main"], marking_location_list
                                ),
                                "animal_id_main": str(row_animal["label_id_main"]),
                                "ref_doc_uuid": str(row_animal["label_id_main"]),
                            }
                        )
                    animals_for_register.append(
                        (
                            row_idx,
                            get_animal_registration_ame_single(**animal),
                        )
                    )
                except Exception as error:
                    progress_callback.emit(f"Строка {row_idx + 1}: {error}")
            else:
                self.table.set_color_to_row(row_idx, (255, 220, 220))
                progress_callback.emit(
                    f"---------------------------------------------------------------------"
                )
        progress_callback.emit(f"Проверка правильности ввода данных в таблицу: конец")
        self.animals_for_register = animals_for_register
        self.status_bar.clearMessage()

    def report_progress(self, info: str):
        self.log_window.write(f"{info}")
