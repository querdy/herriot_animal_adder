import json
import os.path
from datetime import datetime, date
from pathlib import Path

import machineid
from PyQt6.QtCore import QThreadPool, Qt, pyqtSignal
from PyQt6.QtWidgets import QWidget, QStatusBar, QGridLayout, QPushButton, QGroupBox, QLineEdit, QLabel, QPlainTextEdit, \
    QFrame
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, absolute_coordinate
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from app.license_service import load_license, MachineJWT, TokenError, save_license
from app.threads.worker import Worker
from app.vetis.main import Vetis
from app.windows.log import Log


class Service(QWidget):
    def __init__(self, log_window: Log, status_bar: QStatusBar, parent=None):
        super().__init__(parent)

        self.log_window = log_window
        self.status_bar = status_bar
        self.parent = parent
        self.thread_pool = QThreadPool()

        self.layout = QGridLayout()
        self.layout.setVerticalSpacing(10)
        self.layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.setLayout(self.layout)

        # ToDo get individual register excel button
        self.get_individual_register_excel_button = QPushButton()
        self.get_individual_register_excel_button.setText('Получить актуальную excel таблицу (индивидуальный учет)')
        self.get_individual_register_excel_button.setEnabled(False)
        self.get_individual_register_excel_button.clicked.connect(self.get_individual_register_excel)
        self.layout.addWidget(self.get_individual_register_excel_button, 0, 0, 1, 1)

        # ToDo get group register excel button
        self.get_group_register_excel_button = QPushButton()
        self.get_group_register_excel_button.setText('Получить актуальную excel таблицу (групповой учет)')
        self.get_group_register_excel_button.setEnabled(False)
        self.layout.addWidget(self.get_group_register_excel_button, 0, 1, 1, 1)

        # ToDO api settings
        self.api_frame = QGroupBox("Настройки API")
        self.frame_layout = QGridLayout()

        self.initiator_label = QLabel("Логин пользователя:")
        self.initiator_line_edit = QLineEdit()
        self.initiator_line_edit.setMaximumWidth(500)
        self.frame_layout.addWidget(self.initiator_label, 0, 0, 1, 1)
        self.frame_layout.addWidget(self.initiator_line_edit, 0, 1, 1, 7)

        self.issuer_id_label = QLabel("issuer id:")
        self.issuer_id_line_edit = QLineEdit()
        self.issuer_id_line_edit.setMaximumWidth(500)
        self.frame_layout.addWidget(self.issuer_id_label, 1, 0, 1, 1)
        self.frame_layout.addWidget(self.issuer_id_line_edit, 1, 1, 1, 7)

        self.enterprise_login_label = QLabel("Логин API:")
        self.enterprise_login_line_edit = QLineEdit()
        self.enterprise_login_line_edit.setMaximumWidth(500)
        self.frame_layout.addWidget(self.enterprise_login_label, 2, 0, 1, 1)
        self.frame_layout.addWidget(self.enterprise_login_line_edit, 2, 1, 1, 7)

        self.enterprise_password_label = QLabel("Пароль API:")
        self.enterprise_password_line_edit = QLineEdit()
        self.enterprise_password_line_edit.setMaximumWidth(500)
        self.frame_layout.addWidget(self.enterprise_password_label, 3, 0, 1, 1)
        self.frame_layout.addWidget(self.enterprise_password_line_edit, 3, 1, 1, 7)

        self.api_key_label = QLabel("API Key:")
        self.api_key_line_edit = QLineEdit()
        self.api_key_line_edit.setMaximumWidth(500)
        self.frame_layout.addWidget(self.api_key_label, 4, 0, 1, 1)
        self.frame_layout.addWidget(self.api_key_line_edit, 4, 1, 1, 7)

        self.save_api_settings_button = QPushButton("Сохранить")
        self.save_api_settings_button.clicked.connect(self.check_api_data)
        self.frame_layout.addWidget(self.save_api_settings_button, 5, 0, 1, 1)

        self.api_frame.setLayout(self.frame_layout)
        self.layout.addWidget(self.api_frame, 2, 0, 1, 2)

        # ToDo license frame
        self.license_frame = QGroupBox("Лицензия")
        self.license_frame_layout = QGridLayout()

        self.license_key_label = QLabel("Ключ продукта:")
        self.license_key_edit = QLineEdit()
        self.license_frame_layout.addWidget(self.license_key_label, 0, 0, 1, 1)
        self.license_frame_layout.addWidget(self.license_key_edit, 0, 1, 1, 4)

        self.expiry_label = QLabel("Действие до:")
        self.expiry_value_label = QLabel()
        self.license_frame_layout.addWidget(self.expiry_label, 2, 0, 1, 1)
        self.license_frame_layout.addWidget(self.expiry_value_label, 2, 1, 1, 1)

        self.verify_districts_label = QLabel("Территория:")
        self.verify_districts_value = QLineEdit()
        self.verify_districts_value.setReadOnly(True)
        self.license_frame_layout.addWidget(self.verify_districts_label, 2, 2, 1, 1)
        self.license_frame_layout.addWidget(self.verify_districts_value, 2, 3, 1, 2)

        self.machine_id_label = QLabel("ID устройства:")
        self.machine_id_value = QLineEdit(machineid.id())
        self.machine_id_value.setReadOnly(True)
        self.license_frame_layout.addWidget(self.machine_id_label, 1, 0, 1, 1)
        self.license_frame_layout.addWidget(self.machine_id_value, 1, 1, 1, 4)

        self.save_license_key_button = QPushButton("Сохранить")
        self.save_license_key_button.clicked.connect(self.check_license)
        self.license_frame_layout.addWidget(self.save_license_key_button, 3, 0, 1, 1)

        self.license_frame.setLayout(self.license_frame_layout)
        self.layout.addWidget(self.license_frame, 3, 0, 1, 1)

        # ToDo what frame
        self.what_frame = QGroupBox("What is?..")
        self.what_frame_layout = QGridLayout()

        self.what_label_1 = QLabel("tg: @Boyaraa")
        self.what_label_2 = QLabel("email: brainless-slime@ya.ru")
        self.what_label_3 = QLabel("")
        self.what_label_4 = QLabel("")
        self.what_frame_layout.addWidget(self.what_label_1, 0, 0, 1, 1)
        self.what_frame_layout.addWidget(self.what_label_2, 1, 0, 1, 1)
        self.what_frame_layout.addWidget(self.what_label_3, 2, 0, 1, 1)
        self.what_frame_layout.addWidget(self.what_label_4, 3, 0, 1, 1)

        self.what_frame.setLayout(self.what_frame_layout)
        self.layout.addWidget(self.what_frame, 3, 1, 1, 1)

        self.check_license()
        self.check_api_data()

    def report_progress(self, info: str):
        self.log_window.write(f"{info}")

    def check_license(self):
        self.save_license_key_button.setEnabled(False)
        worker = Worker(self.check_license_thread)
        worker.signals.finished.connect(lambda: self.save_license_key_button.setEnabled(True))
        worker.signals.progress.connect(self.report_progress)
        self.thread_pool.start(worker)

    def check_license_thread(self, progress_callback: pyqtSignal, worker_object: Worker):
        districts: dict[str, str] = {
            "a309e4ce-2f36-4106-b1ca-53e0f48a6d95": "Пермь",
            "4f07ade1-1415-44c8-bed9-e851f1ef558d": "Александровск",
            "4ffcde97-05e9-4a6e-bd51-3a984b41b7bd": "Березники",
            "edeaf203-1bc3-4fe1-b3ed-15eb89978783": "Гремячинск",
            "1dc365d3-60b1-41ea-a3b3-1c599024cf30": "Губаха",
            "230c2f4d-fd9d-46fc-8bbd-b8de26810790": "Добрянка",
            "cc8b9eb5-bd4e-4472-8314-f889e8a6679c": "Кизел",
            "d36590ad-0286-44a2-876d-7732deee4234": "Кудымкар",
            "73e5113d-949a-4a9e-8e44-6eae9ef93747": "Кунгур",
            "5d7c95a5-a4d7-4fb4-9774-93e527636a9e": "Лысьва",
            "8b698775-fe5e-4fc0-9f0d-272855d82d15": "Соликамск",
            "06d7a6d6-8f57-4e5a-b698-2bc44c92bb84": "Чайковский",
            "784b1911-182d-476b-a4ad-0f3fa1ef7777": "Чусовой",
            "0810dccb-7114-4d33-9454-50b00433eb1b": "Бардымский район",
            "57e3f364-d709-44a4-a81f-c9ea68778fd0": "Березовский район",
            "76561b13-cc96-478c-9266-bc69ec254776": "Большесосновский район",
            "c3700ef0-0a85-4032-947c-009f956fd754": "Верещагинский район",
            "d336561f-cc98-4742-9f7b-52d99c78463e": "Гайнский район",
            "b6c85a2d-c0ec-4030-a306-4ca7bdcd25f4": "Горнозаводский район",
            "3afa40f2-8169-4006-a5c4-33ace0510d7f": "Еловский район",
            "a80bf180-06e3-4b38-971c-ecef8b417337": "Ильинский район",
            "44166947-11b9-4a1b-a1bf-6e9cba64299d": "Карагайский район",
            "c787b918-d201-408b-abb9-84d53befc6a5": "Кишертский район",
            "6e850977-64c1-49ac-a405-8ff2d77fa219": "Косинский район",
            "a4b9d248-dec4-4686-b4bc-d3b0d8fd9be6": "Кочевский район",
            "e4172d66-d08e-4eda-a274-c47119c30470": "Красновишерский район",
            "1ffa3973-279c-4bcd-a9fc-ece8c7d1039e": "Краснокамский район",
            "6fe17476-90db-4bbf-af4f-e33564751f95": "Кудымкарский район",
            "189875b0-b84b-4980-a125-5a0581f5197e": "Куединский район",
            "1f81a925-75ee-4fa2-87b1-9de26a2813ed": "Кунгурский район",
            "fa41d32f-22b5-4436-bddd-c2ec035377c6": "Нытвенский район",
            "2ddaf37d-e4ea-4748-b6b8-943854f37a0f": "Октябрьский район",
            "5ca82a07-403e-48aa-8fa8-a00277123e46": "Ординский район",
            "60dd742a-3dc9-4ab1-9a22-12c19efdb340": "Осинский район",
            "e7012a7a-e7de-4306-8f52-aabeaf82f178": "Оханский район",
            "2e82fea7-7e6f-4c8a-849f-57140f80c7f3": "Очерский район",
            "7dd380b3-ce33-4280-934f-a4265a07803b": "Пермский район",
            "9eea0415-ab1c-4b49-bd9d-aa04ea23d4e9": "Сивинский район",
            "db4332aa-5444-4c77-a364-541563e0bb1d": "Соликамский район",
            "9460a26c-a2b9-4cb2-9e0e-ddc9022b0ecb": "Суксунский район",
            "a266fff5-5817-4d3b-95dd-c447144f02d1": "Уинский район",
            "7f21b9b5-1f39-44ad-8ce4-6e186e8389ce": "Усольский район",
            "0e9750c7-9f8c-4e23-b996-9a7bff46bb2c": "Частинский район",
            "4f092f1f-8ebf-4ea4-8f01-a75cbcf1d43f": "Чердынский район",
            "75550fdb-56e3-44d5-a4c4-75ab2cb53e83": "Чернушинский район",
            "03547d69-de22-4781-b728-e0823fcdb5f3": "Юрлинский район",
            "f8523e08-3e73-4ba1-b1a2-4bdaf1e8b82f": "Юсьвинский район",
        }
        try:
            new_license_key = self.license_key_edit.text()
            if new_license_key:
                license_key = new_license_key.strip()
            else:
                license_key = load_license()
            license_value = MachineJWT(machineid.id()).decode(license_key)
            if new_license_key:
                save_license(license_key)

            self.expiry_value_label.setText(
                datetime.fromtimestamp(license_value.get("exp")).strftime("%d.%m.%Y %H:%M:%S")
            )
            self.verify_districts_value.setText(
                ", ".join([districts.get(item) for item in license_value.get("district")])
            )
            self.parent.signals.license_checked.emit(True)
            self.parent.signals.allowed_districts_received.emit(license_value.get("district"))
            progress_callback.emit(
                f"Лицензия действительна до "
                f"{datetime.fromtimestamp(license_value.get("exp")).strftime("%d.%m.%Y %H:%M:%S")}, "
                f"Область действия: "
                f"{", ".join([districts.get(item) for item in license_value.get("district")])}"
            )
        except (TokenError, FileNotFoundError) as err:
            self.parent.signals.license_checked.emit(False)
            progress_callback.emit(f"{err}")

    def check_api_data(self):
        self.save_api_settings_button.setEnabled(False)
        worker = Worker(self.check_api_data_thread)
        worker.signals.finished.connect(lambda: self.save_api_settings_button.setEnabled(True))
        worker.signals.progress.connect(self.report_progress)
        self.thread_pool.start(worker)

    def check_api_data_thread(self, progress_callback: pyqtSignal, worker_object: Worker):
        api_data = {}
        try:
            api_data = self.load_api_settings()
        except FileNotFoundError as err:
            # progress_callback.emit(f"{err}")
            pass

        new_api_key = self.api_key_line_edit.text().strip()
        new_initiator = self.initiator_line_edit.text().strip()
        new_issuer_id = self.issuer_id_line_edit.text().strip()
        new_enterprise_login = self.enterprise_login_line_edit.text().strip()
        new_enterprise_password = self.enterprise_password_line_edit.text().strip()

        if new_api_key:
            api_data["api_key"] = new_api_key
        if new_initiator:
            api_data["initiator"] = new_initiator
        if new_issuer_id:
            api_data["issuer_id"] = new_issuer_id
        if new_enterprise_login:
            api_data["enterprise_login"] = new_enterprise_login
        if new_enterprise_password:
            api_data["enterprise_password"] = new_enterprise_password

        if not (all(api_data.values()) and len(api_data.values()) == 5):
            progress_callback.emit(f"Отсутствуют необходимые данные для входа")
            self.parent.signals.vetis_initialized.emit(False)
            self.get_individual_register_excel_button.setEnabled(False)
            return

        Vetis.initialize(
            api_key=api_data["api_key"],
            service_id='herriot.service:1.0',
            issuer_id=api_data["issuer_id"],
            initiator=api_data["initiator"],
            wsdl_dictionary_service='https://api.vetrf.ru/schema/platform/herriot/v1.0b-last/DictionaryService_v1.0.wsdl',
            wsdl_herriot='https://api.vetrf.ru/schema/platform/herriot/v1.0b-last/ams-herriot.service_v1.0.wsdl',
            wsdl_enterprise_service='https://api.vetrf.ru/schema/platform/herriot/v1.0b-last/EnterpriseService_v1.0.wsdl',
            enterprise_login=api_data["enterprise_login"],
            enterprise_password=api_data["enterprise_password"],
        )
        try:
            Vetis.herriot.get_animal_registration_by_guid('47600375-0cda-4122-8fc0-99cd651e90e1')
            response = Vetis.herriot.get_finished_response()
            if response.status == "REJECTED":
                for error in response.errors.error:
                    if error.code == "HRRT112022004":
                        raise ValueError(error._value_1)
            self.parent.signals.vetis_initialized.emit(True)
            self.get_individual_register_excel_button.setEnabled(True)
            self.save_api_settings(
                {
                    "api_key": api_data["api_key"],
                    "issuer_id": api_data["issuer_id"],
                    "initiator": api_data["initiator"],
                    "enterprise_login": api_data["enterprise_login"],
                    "enterprise_password": api_data["enterprise_password"],
                }
            )
        except Exception as err:
            progress_callback.emit(f"Некорректные данные для входа. Проверьте правильность ввода. [{type(err)}: {err}]")
            self.parent.signals.vetis_initialized.emit(False)
            self.get_individual_register_excel_button.setEnabled(False)

    def get_individual_register_excel(self):
        self.get_individual_register_excel_button.setEnabled(False)
        worker = Worker(self.get_individual_register_excel_thread)
        worker.signals.progress.connect(self.report_progress)
        worker.signals.finished.connect(lambda: self.get_individual_register_excel_button.setEnabled(True))
        self.thread_pool.start(worker)

    @staticmethod
    def get_individual_register_excel_thread(progress_callback: pyqtSignal, worker_object: Worker):
        progress_callback.emit(f"Сбор сведений для генерации таблицы")
        animal_species = Vetis.dictionary_service.get_animal_species_list()
        keeping_type = Vetis.dictionary_service.get_animal_keeping_type_list()
        keeping_purposes = Vetis.dictionary_service.get_animal_keeping_purpose_list()
        marking_locations = Vetis.dictionary_service.get_animal_marking_location_list()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Главная'
        wb.create_sheet('animal species')
        wb.create_sheet('validation data')

        ws = wb["animal species"]
        ws.cell(row=1, column=1, value="Вид животного").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Порода").font = Font(bold=True)
        ws.cell(row=1, column=3, value="Вид животного (уникальные)").font = Font(bold=True)
        animal_species_row = 2
        breed_row = 2
        for animal in animal_species:
            ws.cell(row=animal_species_row, column=3, value=animal["name"])
            breeds = Vetis.dictionary_service.get_animal_breed_list(species_guid=animal["guid"])
            for breed in breeds:
                ws.cell(row=breed_row, column=1, value=animal["name"])
                ws.cell(row=breed_row, column=2, value=breed["name"])
                breed_row += 1
            animal_species_row += 1

        ref = f"'animal species'!{absolute_coordinate(f'{ws.cell(row=2, column=2).coordinate}:'
                                                      f'{ws.cell(row=breed_row, column=2).coordinate}')}"
        defn = DefinedName("Порода", attr_text=ref)
        wb.defined_names["Порода"] = defn

        column_count = 1
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('animal species'!$C$2,0,0,COUNTA('animal species'!$C:$C) - 1,1)",
            allow_blank=False
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (вид животного)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Вид животного").font = Font(bold=True, color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 20
        column_count += 1

        dv = DataValidation(
            type="list",
            formula1="=OFFSET('animal species'!$A$2,MATCH('Главная'!$A$2,'animal species'!$A:$A,0)-1,1,COUNTIF('animal species'!$A:$A,'Главная'!$A$2),1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (порода)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Порода").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 20
        column_count += 1

        ws = wb["validation data"]
        ws.cell(row=1, column=2, value="Типы содержания").font = Font(bold=True)
        for idx in range(len(keeping_type)):
            ws.cell(row=idx + 2, column=2, value=keeping_type[idx]["name"])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$B$2,0,0,COUNTA('validation data'!$B:$B) - 1,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (тип содержания)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Тип содержания").font = Font(bold=True, color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 20
        column_count += 1

        ws = wb["Главная"]
        dv = DataValidation(type='date')
        dv.error = 'Вы ввели некорректное значение. Формат ввода: dd.mm.yyyy'
        dv.errorTitle = 'Неверный ввод (дата рождения)'
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Дата рождения").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 15
        column_count += 1

        dv = DataValidation(
            type="list",
            formula1='"самка, самец"',
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (пол)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Пол").font = Font(bold=True, color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 10
        column_count += 1

        dv = DataValidation(
            type="list",
            formula1='"племенное, не племенное, другое"',
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (Племенная ценность)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Плем. ценность").font = Font(bold=True, color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 15
        column_count += 1

        dv = DataValidation(
            type="list",
            formula1='"рождение, импорт, другое"',
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (Причина первичной идентификации)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Причина перв. ид-ии").font = Font(bold=True, color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 20
        column_count += 1

        ws = wb["Главная"]
        dv = DataValidation(type='date')
        dv.error = 'Вы ввели некорректное значение. Формат ввода: dd.mm.yyyy'
        dv.errorTitle = 'Неверный ввод (дата маркирования)'
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Дата маркирования (осн.)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        ws = wb["validation data"]
        ws.cell(row=1, column=3, value="Типы средства маркирования").font = Font(bold=True)
        marking_means = ["Бирка", "Вживляемый микрочип", "Болюс", "Кольцо", "Электронное кольцо", "Ошейник",
                         "Электронный ошейник", "Крыло-метка", "Электронное крыло-метка", "Электронная метка", "Табло"]
        for idx in range(len(marking_means)):
            ws.cell(row=idx + 2, column=4, value=marking_means[idx])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$D$2,0,0,COUNTA('validation data'!$D:$D) - 0,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (Вид средства маркирования)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Вид ср-ва марк. (осн.)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 21
        column_count += 1

        ws = wb["validation data"]
        ws.cell(row=1, column=3, value="Marking locations").font = Font(bold=True)
        for idx in range(len(marking_locations)):
            ws.cell(row=idx + 2, column=3, value=marking_locations[idx]["name"])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$C$2,0,0,COUNTA('validation data'!$C:$C) - 0,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (место нанесения средства маркирования)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Место нанесения СМ (осн.)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 26
        column_count += 1

        ws.cell(row=1, column=column_count, value="Инд. номер (осн.)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 19
        column_count += 1

        ws = wb["Главная"]
        dv = DataValidation(type='date')
        dv.error = 'Вы ввели некорректное значение. Формат ввода: dd.mm.yyyy'
        dv.errorTitle = 'Неверный ввод (дата маркирования)'
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Дата маркирования (доп. 1)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 27
        column_count += 1

        ws = wb["validation data"]
        ws.cell(row=1, column=3, value="Типы средства маркирования").font = Font(bold=True)
        marking_means = ["Бирка", "Вживляемый микрочип", "Болюс", "Кольцо", "Электронное кольцо", "Ошейник",
                         "Электронный ошейник", "Крыло-метка", "Электронное крыло-метка", "Электронная метка", "Табло"]
        for idx in range(len(marking_means)):
            ws.cell(row=idx + 2, column=4, value=marking_means[idx])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$D$2,0,0,COUNTA('validation data'!$D:$D) - 0,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (Вид средства маркирования)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Вид ср-ва марк. (доп. 1)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 23
        column_count += 1

        ws = wb["validation data"]
        ws.cell(row=1, column=3, value="Marking locations").font = Font(bold=True)
        for idx in range(len(marking_locations)):
            ws.cell(row=idx + 2, column=3, value=marking_locations[idx]["name"])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$C$2,0,0,COUNTA('validation data'!$C:$C) - 0,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (место нанесения средства маркирования)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Место нанесения СМ (доп. 1)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 28
        column_count += 1

        ws.cell(row=1, column=column_count, value="Инд. номер (доп. 1)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 21
        column_count += 1

        ws = wb["Главная"]
        dv = DataValidation(type='date')
        dv.error = 'Вы ввели некорректное значение. Формат ввода: dd.mm.yyyy'
        dv.errorTitle = 'Неверный ввод (дата маркирования)'
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Дата маркирования (доп. 2)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 27
        column_count += 1

        ws = wb["validation data"]
        ws.cell(row=1, column=3, value="Типы средства маркирования").font = Font(bold=True)
        marking_means = ["Бирка", "Вживляемый микрочип", "Болюс", "Кольцо", "Электронное кольцо", "Ошейник",
                         "Электронный ошейник", "Крыло-метка", "Электронное крыло-метка", "Электронная метка", "Табло"]
        for idx in range(len(marking_means)):
            ws.cell(row=idx + 2, column=4, value=marking_means[idx])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$D$2,0,0,COUNTA('validation data'!$D:$D) - 0,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (Вид средства маркирования)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Вид ср-ва марк. (доп. 2)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 23
        column_count += 1

        ws = wb["validation data"]
        ws.cell(row=1, column=3, value="Marking locations").font = Font(bold=True)
        for idx in range(len(marking_locations)):
            ws.cell(row=idx + 2, column=3, value=marking_locations[idx]["name"])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$C$2,0,0,COUNTA('validation data'!$C:$C) - 0,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (место нанесения средства маркирования)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Место нанесения СМ (доп. 2)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 28
        column_count += 1

        ws.cell(row=1, column=column_count, value="Инд. номер (доп. 2)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 21
        column_count += 1

        ws.cell(row=1, column=column_count, value="Окрас").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 15
        column_count += 1
        ws.cell(row=1, column=column_count, value="Кличка").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 15
        column_count += 1
        # ws.cell(row=1, column=column_count, value="ИНН орг-ии, нанесшей марк.").font = Font(bold=True)
        # ws.column_dimensions[get_column_letter(column_count)].width = 25
        # column_count += 1
        ws.cell(row=1, column=column_count, value="Место содержания (ПО)").font = Font(bold=True, color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        ws.cell(row=1, column=column_count, value="ИНН владельца ПО места содержания").font = Font(bold=True,
                                                                                                   color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 40
        column_count += 1

        ws.cell(row=1, column=column_count, value="Место рождения (ПО)").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        ws.cell(row=1, column=column_count, value="ИНН владельца ПО места рождения").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 40
        column_count += 1

        ws = wb["validation data"]
        ws["A1"] = "цели содержания"
        ws["A1"].font = Font(bold=True)

        for idx in range(len(keeping_purposes)):
            ws.cell(row=idx + 2, column=1, value=keeping_purposes[idx]["name"])
        dv = DataValidation(
            type="list",
            formula1="=OFFSET('validation data'!$A$2,0,0,COUNTA('validation data'!$A:$A) - 1,1)",
            allow_blank=True
        )
        dv.error = 'Вы ввели значение не из списка'
        dv.errorTitle = 'Неверный ввод (цель содержания)'
        dv.showErrorMessage = True

        ws = wb["Главная"]
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Цель содержания 1").font = Font(bold=True, color="008000")
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Цель содержания 2").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Цель содержания 3").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Цель содержания 4").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        dv.add(ws.cell(row=2, column=column_count).coordinate)
        ws.cell(row=1, column=column_count, value="Цель содержания 5").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 25
        column_count += 1

        ws.cell(row=1, column=column_count, value="Родитель 1").font = Font(bold=True)
        ws.cell(row=1, column=column_count).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(column_count)].width = 20
        column_count += 1

        desktop_path = Path(os.path.expanduser("~/Desktop"))
        table_name = "Таблица индивидуального учета животных.xlsx"
        try:
            wb.save(f"{desktop_path}/{table_name}")
        except PermissionError:
            processes = ('excel.exe', 'EXCEL.EXE', 'excel.EXE', 'EXCEL.exe')
            for process in processes:
                os.system("chcp 65001 > nul")
                os.system(f"taskkill /f /im {process}")
            wb.save(f"{desktop_path}/{table_name}")
        progress_callback.emit(f"Файл `{table_name}` сохранен на рабочий стол ({desktop_path})")

    @staticmethod
    def save_api_settings(data: dict) -> bool:
        try:
            json.dump(data, open('api', 'w'), indent=4)
            return True
        except Exception:
            return False

    @staticmethod
    def load_api_settings() -> dict:
        try:
            return json.load(open('api', 'r'))
        except FileNotFoundError:
            raise FileNotFoundError(f"Файл с данными для входа не обнаружен")
